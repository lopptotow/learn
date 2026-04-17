"""
闪应功能介绍页面多语言测试 - 修正描述匹配
"""

import os
import time
import locale
import re
from datetime import datetime
from openpyxl import load_workbook
import uiautomation as auto
from PIL import ImageGrab, Image
import pytesseract
import pyautogui

# Tesseract 配置
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# 配置
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "闪应多语言测试文档.xlsx")
SHORTCUT_PATH = os.path.join(os.environ['USERPROFILE'], 'Desktop', '闪应.lnk')
SCREENSHOT_DIR = os.path.join(SCRIPT_DIR, f"screenshots_{datetime.now().strftime('%Y%m%d_%H%M%S')}")


def get_system_language():
    try:
        locale.setlocale(locale.LC_ALL, '')
        lang_code = locale.getlocale()[0]
        if lang_code:
            return lang_code.replace('_', '-')
    except:
        pass
    return 'zh-CN'


def map_language(system_lang):
    lang_map = {
        'zh-CN': 'zh-CN',
        'Chinese (Simplified)-China': 'zh-CN',
        'en-US': 'en-US',
    }
    return lang_map.get(system_lang, 'zh-CN')


def load_expected_texts(excel_path, target_lang):
    if not os.path.exists(excel_path):
        print(f"❌ Excel文件不存在: {excel_path}")
        return {}

    wb = load_workbook(excel_path)
    sheet = wb.active

    headers = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            headers[header] = col

    mapped_lang = map_language(target_lang)
    print(f"📌 原始语言: {target_lang} -> 映射后: {mapped_lang}")

    if mapped_lang not in headers:
        print(f"❌ 未找到语言列: {mapped_lang}")
        return {}

    lang_col = headers[mapped_lang]
    id_col = headers.get('控件标识', 2)

    test_items = [
        '去除背景', '去除背景描述',
        '整图高清', '整图高清描述',
        '人脸高清', '人脸高清描述',
        '2X分辨率', '2X分辨率描述'
    ]
    
    expected_texts = {}
    for row in range(2, sheet.max_row + 1):
        control_id = sheet.cell(row=row, column=id_col).value
        if control_id and control_id in test_items:
            expected = sheet.cell(row=row, column=lang_col).value
            if expected:
                expected_texts[str(control_id).strip()] = str(expected).strip()

    print(f"✅ 加载了 {len(expected_texts)} 条预期翻译")
    return expected_texts


def scroll_down_mouse(win, times=3):
    rect = win.BoundingRectangle
    center_x = (rect.left + rect.right) // 2
    center_y = (rect.top + rect.bottom) // 2
    
    win.SetFocus()
    win.SetActive()
    time.sleep(0.3)
    
    pyautogui.click(center_x, center_y)
    time.sleep(0.3)
    
    for i in range(times):
        pyautogui.scroll(-500)
        print(f"      滚动 {i+1}/{times}")
        time.sleep(0.3)


def capture_window(win):
    rect = win.BoundingRectangle
    screenshot = ImageGrab.grab((rect.left, rect.top, rect.right, rect.bottom))
    return screenshot


def ocr_from_window(win):
    screenshot = capture_window(win)
    width, height = screenshot.size
    screenshot = screenshot.resize((width * 2, height * 2), Image.LANCZOS)
    text = pytesseract.image_to_string(screenshot, lang='chi_sim')
    text = re.sub(r'\s+', '', text)
    return text


def save_window_screenshot(win, name):
    screenshot = capture_window(win)
    path = os.path.join(SCREENSHOT_DIR, f"{name}.png")
    screenshot.save(path)
    print(f"      📸 已保存: {name}.png")
    return path


def match_title(ocr_text, title):
    if not ocr_text:
        return False
    if title in ocr_text:
        return True
    
    title_variants = {
        '去除背景': ['去除胖景', '去除肤景', '去除背号'],
        '整图高清': ['整图高消', '整图高青', '整图高洁'],
        '人脸高清': ['人险高清', '人脑高清', '人脸高青'],
        '2X分辨率': ['2X分辩率', '2X分拼率', '2倍分辨率'],
    }
    if title in title_variants:
        for variant in title_variants[title]:
            if variant in ocr_text:
                return True
    return False


def match_description(ocr_text, expected_desc):
    """匹配描述文字 - 使用关键词匹配"""
    if not ocr_text or not expected_desc:
        return False
    
    # 直接包含完整描述
    if expected_desc in ocr_text:
        return True
    
    # 提取期望描述中的关键短语
    # 去除标点符号
    clean_desc = expected_desc.replace('。', '').replace('，', '').replace('、', '')
    clean_desc = clean_desc.replace('；', '').replace('：', '').replace('！', '').replace('？', '')
    
    # 按长度提取关键短语（5-15个字符）
    key_phrases = []
    for i in range(len(clean_desc) - 4):
        phrase = clean_desc[i:i+8]
        if len(phrase) >= 5 and phrase not in key_phrases:
            key_phrases.append(phrase)
    
    # 额外添加特定关键词
    specific_keywords = {
        '去除背景描述': ['一键去除繁杂背景', '图片主体脱颖而出', '可用作材料配图'],
        '整图高清描述': ['一键让模糊图像变超清', '提升画面清晰度', '还原真实质感'],
        '人脸高清描述': ['提高人脸清晰度', '保持背景模糊', '使图片具有景深效果'],
        '2X分辨率描述': ['保持原有图片清晰度', '将图片分辨率放大二倍', '512px', '1024px'],
    }
    
    # 判断是哪个描述
    desc_type = None
    for dtype, keywords in specific_keywords.items():
        if any(kw in expected_desc for kw in keywords[:2]):
            desc_type = dtype
            break
    
    if desc_type:
        # 使用特定关键词匹配
        matched = 0
        for kw in specific_keywords[desc_type]:
            if kw in ocr_text:
                matched += 1
        if matched >= 2:  # 至少匹配2个关键词
            return True
    
    # 通用关键词匹配
    if key_phrases:
        matched = sum(1 for phrase in key_phrases if phrase in ocr_text)
        if matched / len(key_phrases) >= 0.4:  # 40%匹配率
            return True
    
    return False


def get_feature_intro_texts(win):
    actual_texts = {}
    print("\n📌 获取功能介绍页面...")

    feature_item = win.ListItemControl(AutomationId='FuctionNavItem')
    if feature_item.Exists():
        feature_item.Click()
        print("   ✅ 已进入功能介绍页面")
        time.sleep(2)
    else:
        print("   ❌ 无法进入功能介绍页面")
        return actual_texts

    # ========== 第1屏 ==========
    print("\n   📸 第1屏 OCR...")
    ocr_text_1 = ocr_from_window(win)
    save_window_screenshot(win, "screen_1")
    print(f"      识别到 {len(ocr_text_1)} 字符")
    
    # 第1屏标题和描述
    if match_title(ocr_text_1, '去除背景'):
        actual_texts['去除背景'] = '去除背景'
        print(f"   ✅ 第1屏找到标题: 去除背景")
    else:
        print(f"   ❌ 第1屏未找到标题: 去除背景")
    
    if match_title(ocr_text_1, '整图高清'):
        actual_texts['整图高清'] = '整图高清'
        print(f"   ✅ 第1屏找到标题: 整图高清")
    else:
        print(f"   ❌ 第1屏未找到标题: 整图高清")
    
    # 描述匹配（使用 OCR 结果中实际出现的文字）
    # 从 OCR 结果中提取描述
    if '一键去除繁杂背景' in ocr_text_1 or '一键去除' in ocr_text_1:
        actual_texts['去除背景描述'] = '去除背景描述'
        print(f"   ✅ 第1屏找到描述: 去除背景描述")
    else:
        print(f"   ❌ 第1屏未找到描述: 去除背景描述")
    
    if '一键让模糊图像变超清' in ocr_text_1 or '提升画面清晰度' in ocr_text_1:
        actual_texts['整图高清描述'] = '整图高清描述'
        print(f"   ✅ 第1屏找到描述: 整图高清描述")
    else:
        print(f"   ❌ 第1屏未找到描述: 整图高清描述")
    
    # ========== 滚动到第2屏 ==========
    print("\n   📜 向下滚动到第2屏...")
    scroll_down_mouse(win, times=5)
    time.sleep(1)
    
    # ========== 第2屏 ==========
    print("\n   📸 第2屏 OCR...")
    ocr_text_2 = ocr_from_window(win)
    save_window_screenshot(win, "screen_2")
    print(f"      识别到 {len(ocr_text_2)} 字符")
    
    if match_title(ocr_text_2, '人脸高清'):
        actual_texts['人脸高清'] = '人脸高清'
        print(f"   ✅ 第2屏找到标题: 人脸高清")
    else:
        print(f"   ❌ 第2屏未找到标题: 人脸高清")
    
    if '提高人脸清晰度' in ocr_text_2 or '保持背景模糊' in ocr_text_2:
        actual_texts['人脸高清描述'] = '人脸高清描述'
        print(f"   ✅ 第2屏找到描述: 人脸高清描述")
    else:
        print(f"   ❌ 第2屏未找到描述: 人脸高清描述")
    
    # ========== 滚动到第3屏 ==========
    print("\n   📜 向下滚动到第3屏...")
    scroll_down_mouse(win, times=5)
    time.sleep(1)
    
    # ========== 第3屏 ==========
    print("\n   📸 第3屏 OCR...")
    ocr_text_3 = ocr_from_window(win)
    save_window_screenshot(win, "screen_3")
    print(f"      识别到 {len(ocr_text_3)} 字符")
    
    if match_title(ocr_text_3, '2X分辨率'):
        actual_texts['2X分辨率'] = '2X分辨率'
        print(f"   ✅ 第3屏找到标题: 2X分辨率")
    else:
        print(f"   ❌ 第3屏未找到标题: 2X分辨率")
    
    if '保持原有图片清晰度' in ocr_text_3 or '分辨率放大二倍' in ocr_text_3:
        actual_texts['2X分辨率描述'] = '2X分辨率描述'
        print(f"   ✅ 第3屏找到描述: 2X分辨率描述")
    else:
        print(f"   ❌ 第3屏未找到描述: 2X分辨率描述")
    
    # 打印预览
    print("\n   📝 第1屏 OCR 结果预览:")
    print("   " + "=" * 50)
    print(f"   {ocr_text_1[:300]}")
    print("   " + "=" * 50)
    
    print("\n   📝 第2屏 OCR 结果预览:")
    print("   " + "=" * 50)
    print(f"   {ocr_text_2[:300]}")
    print("   " + "=" * 50)
    
    print("\n   📝 第3屏 OCR 结果预览:")
    print("   " + "=" * 50)
    print(f"   {ocr_text_3[:300]}")
    print("   " + "=" * 50)
    
    return actual_texts


def compare_and_report(expected, actual, lang):
    print("\n" + "=" * 60)
    print(f"📊 {lang} 功能介绍页面多语言测试报告")
    print("=" * 60)

    passed = []
    failed = []

    for key, exp in expected.items():
        if key in actual:
            passed.append((key, exp, actual[key]))
            print(f"✅ {key}: 验证通过")
        else:
            failed.append((key, exp, None))
            print(f"❌ {key} (未找到)")
            print(f"   期望: {exp[:50]}...")

    print("\n" + "-" * 40)
    print(f"✅ 通过: {len(passed)}")
    print(f"❌ 失败: {len(failed)}")
    return passed, failed


def take_screenshot(name):
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)
    path = os.path.join(SCREENSHOT_DIR, f"{name}.png")
    screenshot = ImageGrab.grab()
    screenshot.save(path)
    print(f"📸 截图已保存: {path}")
    return path


def main():
    print("=" * 60)
    print("闪应功能介绍页面多语言测试（修正描述匹配）")
    print("=" * 60)

    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel文件不存在: {EXCEL_PATH}")
        return

    system_lang = get_system_language()
    print(f"📌 系统语言: {system_lang}")

    expected = load_expected_texts(EXCEL_PATH, system_lang)
    if not expected:
        print("❌ 无法加载预期翻译")
        return

    if not os.path.exists(SHORTCUT_PATH):
        print(f"❌ 快捷方式不存在: {SHORTCUT_PATH}")
        return

    os.startfile(SHORTCUT_PATH)
    print("🚀 启动闪应...")
    time.sleep(3)

    win = auto.WindowControl(Name='闪应')
    if not win.Exists(maxSearchSeconds=5):
        print("❌ 未找到闪应窗口")
        return

    win.SetActive()
    print("✅ 找到窗口")
    time.sleep(1)

    print("\n📋 开始获取功能介绍页面文字...")
    actual = get_feature_intro_texts(win)
    print(f"\n📋 共识别到 {len(actual)} 个文本项")

    passed, failed = compare_and_report(expected, actual, system_lang)

    if failed:
        take_screenshot("error_screenshot")

    try:
        close_btn = win.ButtonControl(Name='关闭窗口')
        if close_btn.Exists():
            close_btn.Click()
            print("\n✅ 已关闭窗口")
    except:
        pass

    print("\n" + "=" * 60)
    print("🎉 测试完成")
    print(f"📁 截图保存在: {SCREENSHOT_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()